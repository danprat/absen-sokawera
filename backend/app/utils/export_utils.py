"""
Export Utilities for PDF and Excel generation
"""
import io
from datetime import datetime
from typing import Optional

# PDF imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_pdf(
    title: str,
    subtitle: str,
    headers: list[str],
    data: list[list],
    logo_path: Optional[str] = None,
    orientation: str = "portrait"
) -> bytes:
    """
    Generate a professional PDF report with table data.

    Args:
        title: Main title of the report
        subtitle: Subtitle (e.g., period, filters)
        headers: List of column headers
        data: List of rows, each row is a list of cell values
        logo_path: Optional path to logo image
        orientation: "portrait" or "landscape"

    Returns:
        PDF file as bytes
    """
    buffer = io.BytesIO()

    # Page setup
    page_size = landscape(A4) if orientation == "landscape" else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=2*cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=6,
        textColor=colors.HexColor('#1e3a5f')
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#666666')
    )

    # Add logo if provided
    if logo_path:
        try:
            img = Image(logo_path, width=2*cm, height=2*cm)
            img.hAlign = 'CENTER'
            elements.append(img)
            elements.append(Spacer(1, 0.3*cm))
        except Exception:
            pass  # Skip logo if file not found

    # Add title and subtitle
    elements.append(Paragraph(title.upper(), title_style))
    elements.append(Paragraph(subtitle, subtitle_style))

    # Prepare table data with headers
    table_data = [headers] + data

    # Calculate column widths based on content
    page_width = page_size[0] - 3*cm  # Account for margins
    num_cols = len(headers)
    col_width = page_width / num_cols
    col_widths = [col_width] * num_cols

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table styling
    style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2563eb')),
    ])

    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc'))
        else:
            style.add('BACKGROUND', (0, i), (-1, i), colors.white)

    table.setStyle(style)
    elements.append(table)

    # Footer with generation date
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#9ca3af')
    )
    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    elements.append(Paragraph(f"Dicetak: {generated_at}", footer_style))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel(
    title: str,
    subtitle: str,
    headers: list[str],
    data: list[list],
    sheet_name: str = "Data"
) -> bytes:
    """
    Generate a professional Excel report with styled table.

    Args:
        title: Main title of the report
        subtitle: Subtitle (e.g., period, filters)
        headers: List of column headers
        data: List of rows, each row is a list of cell values
        sheet_name: Name of the worksheet

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    title_font = Font(name='Arial', size=14, bold=True, color='1e3a5f')
    subtitle_font = Font(name='Arial', size=10, color='666666')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    data_font = Font(name='Arial', size=9)

    # Alternating row fills
    even_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Border style
    thin_border = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb')
    )

    # Center alignment
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title.upper())
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 25

    # Row 2: Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center_align
    ws.row_dimensions[2].height = 20

    # Row 3: Empty spacer
    ws.row_dimensions[3].height = 10

    # Row 4: Headers
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 25

    # Data rows starting from row 5
    data_start_row = 5
    for row_idx, row_data in enumerate(data):
        excel_row = data_start_row + row_idx
        fill = even_fill if row_idx % 2 == 0 else odd_fill

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)

        # Calculate max width based on header and data
        max_length = len(str(header))
        for row_data in data:
            if col_idx <= len(row_data):
                cell_value = str(row_data[col_idx - 1]) if row_data[col_idx - 1] is not None else ""
                max_length = max(max_length, len(cell_value))

        # Set width with some padding (min 10, max 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # Add generation timestamp at the bottom
    last_row = data_start_row + len(data) + 1
    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    footer_cell = ws.cell(row=last_row, column=len(headers), value=f"Dicetak: {generated_at}")
    footer_cell.font = Font(name='Arial', size=8, color='9ca3af')
    footer_cell.alignment = Alignment(horizontal='right')

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_csv(headers: list[str], data: list[list]) -> str:
    """
    Generate CSV content as string.

    Args:
        headers: List of column headers
        data: List of rows

    Returns:
        CSV content as string
    """
    import csv

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)

    output.seek(0)
    return output.getvalue()
